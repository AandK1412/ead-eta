"""Official USCIS quarterly volume data — the unbiased counterweight to Reddit.

Two reports, both plain XLSX on www.uscis.gov (no Cloudflare, unlike `egov`):

* **All USCIS Application and Petition Form Types** — receipts, approvals,
  denials, completions, pending and a published processing time, broken out
  into four I-765 categories.
* **Net Backlog and Frontlog** — how many pending cases are already past
  USCIS's own target, in six I-765 categories.

Why this matters: USCIS's published processing time is a median of *completed*
cases, so it carries exactly the survivorship bias this project exists to
correct. Dividing the pending backlog by quarterly throughput (Little's Law)
gives an independent estimate of how long the queue really is — and for the
large categories it comes out several times higher.

Little's Law caveat, carried through to the UI: W = L/λ is the mean time in
system for *everything currently queued*, including long-stuck cases, and
adjudication is not strictly FIFO. Treat it as backlog depth — an upper bound on
a new filer's expected wait, not a median.
"""

from __future__ import annotations

import io
import re

import requests

DATA_PAGE = "https://www.uscis.gov/tools/reports-and-studies/immigration-and-citizenship-data"
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36")

# Fallbacks if link discovery fails (e.g. the page markup changes).
FALLBACK = {
    "forms": "https://www.uscis.gov/sites/default/files/document/reports/quarterly_all_forms_fy2026_q1_v1.xlsx",
    "backlog": "https://www.uscis.gov/sites/default/files/document/reports/net_backlog_frontlog_fy2026_q1_v1.xlsx",
}

# USCIS's coarse buckets -> this project's category ids. "All Other" is the
# catch-all that actually contains OPT, H-4 and L-2, which is why the official
# numbers can only ever be a coarse reference line.
BUCKET_MAP = {
    "asylum": ["c08_asylum"],
    "adjustment of status": ["c09_aos"],
    "daca": ["c33_daca"],
    "parolees": ["c11_parole"],
    "premium processed": [],
    "all other": ["c03a_preopt", "c03b_opt", "c03b_stem", "c26_h4",
                  "a17_a18_l2e2", "a12_c19_tps", "a05_asylee", "other"],
}


def _bucket_key(title: str) -> str | None:
    """'Application for Employment Authorization (Asylum)' -> 'asylum'."""
    m = re.search(r"\(([^)]+)\)\s*$", (title or "").strip())
    return m.group(1).strip().lower() if m else None


def _get(url: str) -> bytes:
    r = requests.get(url, headers={"User-Agent": UA}, timeout=90)
    r.raise_for_status()
    return r.content


def discover_urls() -> dict[str, str]:
    """Find the current quarter's files so this keeps working after each release."""
    urls = dict(FALLBACK)
    try:
        html = requests.get(DATA_PAGE, headers={"User-Agent": UA}, timeout=60).text
        for key, pat in (("forms", r"quarterly_all_forms_fy\d{4}_q\d_v\d+\.xlsx"),
                         ("backlog", r"net_backlog_frontlog_fy\d{4}_q\d_v\d+\.xlsx")):
            if (m := re.search(rf"https://[^\"']*{pat}", html)):
                urls[key] = m.group(0)
    except requests.RequestException as exc:
        print(f"  ! USCIS data page: {exc} — using fallback URLs")
    return urls


def _rows(blob: bytes, want_cols: int):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(values_only=True):
        if row and str(row[0]).strip() == "I-765" and len(row) >= want_cols:
            yield row
    wb.close()


def fetch() -> dict | None:
    """Return the official I-765 picture, or None if the reports are unreachable."""
    urls = discover_urls()

    try:
        forms_blob = _get(urls["forms"])
    except requests.RequestException as exc:
        print(f"  ! USCIS quarterly forms report: {exc}")
        return None

    # Columns: form, title, received, approved, denied, completions, pending, processing_time
    buckets: dict[str, dict] = {}
    period = None
    for row in _rows(forms_blob, 8):
        key = _bucket_key(row[1])
        if not key:
            continue
        try:
            received = float(row[2]); completions = float(row[5]); pending = float(row[6])
        except (TypeError, ValueError):
            continue
        try:
            published = float(row[7])
        except (TypeError, ValueError):
            published = None

        implied = round(pending / completions * 3, 1) if completions > 0 else None
        buckets[key] = {
            "bucket": key,
            "label": str(row[1]),
            "received": int(received),
            "approved": int(row[3]) if isinstance(row[3], (int, float)) else None,
            "completions": int(completions),
            "pending": int(pending),
            "published_months": published,
            "implied_months": implied,
            # Near 1.0 means arrivals ≈ departures, where Little's Law is most
            # trustworthy; far above 1.0 means the queue is growing.
            "inflow_ratio": round(received / completions, 2) if completions > 0 else None,
            "categories": BUCKET_MAP.get(key, []),
        }

    if not buckets:
        print("  ! no I-765 rows found in the quarterly report")
        return None

    # Net backlog: pending cases already past USCIS's own target.
    try:
        for row in _rows(_get(urls["backlog"]), 3):
            key = _bucket_key(row[1])
            if key and key in buckets and isinstance(row[2], (int, float)):
                b = buckets[key]
                b["net_backlog"] = int(row[2])
                b["share_past_target"] = (round(row[2] / b["pending"], 3)
                                          if b["pending"] else None)
    except requests.RequestException as exc:
        print(f"  ! USCIS net backlog report: {exc}")

    # Reporting period, for honest labelling of how stale the quarter is.
    m = re.search(r"fy(\d{4})_q(\d)", urls["forms"])
    if m:
        period = f"FY{m.group(1)} Q{m.group(2)}"

    return {
        "period": period,
        "source_urls": urls,
        "buckets": list(buckets.values()),
    }


if __name__ == "__main__":  # quick manual check
    import json
    print(json.dumps(fetch(), indent=1))
